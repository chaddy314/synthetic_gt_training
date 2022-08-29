import glob
import sys
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
import re
import matplotlib.pyplot as plt
import numpy as np
from natsort import natsorted, ns


class Confusion:
    def __init__(self, gt, pred, count):
        self.gt = str(gt)
        self.pred = str(pred)
        self.count = count


class ConfusionCategory:
    def __init__(self, name):
        self.name = name
        self.count = 0
        self.confusions = []

    def add(self, confusion):
        self.confusions.append(confusion)
        self.count += confusion.count


similar_chars = list("Q00OOooecçB8ßzZ2qg9ßGuvkxKX")


def cat_as_similar(confusion):
    return confusion.gt.strip() in similar_chars and confusion.pred.strip() in similar_chars


resolution_gt = list("mhMNrnwvWPG")
resolution_pred = list("nbNMmvwVFC")


def cat_as_resolution(confusion):
    for char in confusion.gt.strip():
        if char not in resolution_gt:
            return False
    for char in confusion.pred.strip():
        if char not in resolution_pred:
            return False
    return True
    # return confusion.gt.strip() in resolution_gt and confusion.pred.strip() in resolution_pred


def cat_as_letter_case(confusion):
    return confusion.gt.lower().strip() == confusion.pred.lower().strip()


unknown_characters = list("øφû⅙ʿʾāâεēæ")


def cat_as_unknown_character(confusion):
    return confusion.gt.strip() in unknown_characters


german = list("äüöÄÖÜ")


def cat_as_umlaut(confusion):
    return confusion.gt.strip() in german or confusion.pred.strip() in german


def cat_as_whitespace(confusion):
    return (confusion.gt == "" and confusion.pred == " ") or (confusion.gt == " " and confusion.pred == "")


super_script = list("ᴬᴮᶜᴰᴱᶠᴳᴴᴵᴶᴷᴸᴹᴺᴼᴾᴿˢᵀᵁⱽᵂˣʸᶻᵃᵇᶜᵈᵉᶠᵍʰᶦʲᵏˡᵐⁿᵒᵖ۹ʳˢᵗᵘᵛʷˣʸᶻ⁰¹²³⁴⁵⁶⁷⁸⁹⁺⁻⁼⁽⁾")
sub_script = list("ₐ₈ₑₕᵢⱼₖₗₘₙₒₚᵣₛₜᵤᵥₓᵧₐ♭꜀ₑ₉ₕᵢⱼₖₗₘₙₒₚᵣₛₜᵤᵥₓᵧ₂₀₁₂₃₄₅₆₇₈₉₊₋₌₍₎")


def cat_as_sub_super(confusion):
    return any(item in list(confusion.gt.strip()) for item in super_script) or any(item in list(confusion.gt.strip()) for item in sub_script)


vertical_lines = list("iITtl1|/![]()f")


def cat_as_vert_line(confusion):
    for char in confusion.gt.strip():
        if char not in vertical_lines:
            return False
    for char in confusion.pred.strip():
        if char not in vertical_lines:
            return False
    # return confusion.gt.strip() in vertical_lines and confusion.pred.strip() in vertical_lines
    return True


def cat_as_capitals(confusion):
    return len(confusion.gt.strip()) > 1 and confusion.gt.strip().isupper()


kerning_gt = "rn ch ck fl m".split(" ")
kerning_pred = list("mdßrn")


def cat_as_kerning(confusion):
    return confusion.gt.strip() in kerning_gt and confusion.pred.strip() in kerning_pred


upper_quotations = list("\"\'`´”‟“＂〞‘‛’")


def cat_as_upper_quote(confusion):
    return confusion.gt.strip() in upper_quotations and confusion.pred.strip() in upper_quotations


punctuations = list(",.;:-")


def cat_as_punctuation(confusion):
    return confusion.gt.strip() in punctuations and confusion.pred.strip() in punctuations


def main():
    global_model_evaluations = {}
    line_evaluations = {}
    eval_folder = sys.argv[1]
    eval_wb_paths = glob.glob(eval_folder + "*.xlsx")
    eval_wb_paths = natsorted(eval_wb_paths)
    char_count_map = {}
    for wb_path in eval_wb_paths:

        eval_categories = init_categories()
        wb = openpyxl.load_workbook(wb_path)
        print(wb.sheetnames)
        global_eval_sheet = wb['evaluation - global']

        for row in global_eval_sheet.iter_rows(min_row=2):
            gt, pred, count, percent = row
            if max(len(gt.value)-2, len(pred.value)-2) <= 0:
                print("!!!!WARNING!!!!!")
            error_value = count.value * max(len(gt.value[1:-1]), len(pred.value[1:-1]))
            #print("[", str(count.value), ",", error_value, "]")
            confusion = Confusion(str(gt.value), str(pred.value), error_value)
            if len(confusion.gt) == 1 or len(confusion.pred) == 1:
                pass    # debug
            else:
                confusion.gt = confusion.gt[1:-1]
                confusion.pred = confusion.pred[1:-1]
                categorize_confusion(eval_categories, confusion)
        global_model_evaluations[wb_path.split("/")[-1]] = eval_categories

        line_sheet = wb['evaluation - per line']
        print("\n\n-----", wb_path.split("/")[-1], "-----")
        char_count_map[wb_path.split("/")[-1]] = get_char_count(line_sheet)


#        model_line_evaluations = {}
#        model_line_evaluations[0] = init_categories()
#        model_line_evaluations[1] = init_categories()
#        model_line_evaluations[2] = init_categories()
#        model_line_evaluations[3] = init_categories()
#        model_line_evaluations[4] = init_categories()

#        for row in line_sheet.iter_rows(min_row=2):
#            gt_file, gt, pred, line_len, err, cer, rel_err, sync_err, confusions = row
#            if err.value != 0:
#                line_confusions = parse_confusions(confusions)
#                line_cat = int(gt_file.value.split("/")[2])
#                for line_confusion in line_confusions:
#                    categorize_confusion(model_line_evaluations[(line_cat - 1) // 10], line_confusion)
#        line_evaluations[wb_path.split("/")[-1]] = model_line_evaluations
#
#    for evaluation in line_evaluations.keys():
#        print('\n\033[1m' + evaluation + '\033[0m\n')
#        for book_summary in line_evaluations[evaluation].keys():
#            print(book_summary)
#            labels = list(line_evaluations[evaluation][book_summary].keys())
#            labels.sort()
#            for dunno in labels:
#                print("\t", line_evaluations[evaluation][book_summary][dunno].name + ": " + str(line_evaluations[evaluation][book_summary][dunno].count))
#            print("---------------------------------")

    plot_global(global_model_evaluations, char_count_map)
    # cat_unknown_confusion.confusions.sort(key=lambda x: x.count, reverse=False)
    # for confusion in cat_unknown_confusion.confusions:
    #      print("{", confusion.gt, "} - {", confusion.pred, "} - ", confusion.count)


def plot_global(global_register, char_count_map):
    labels = list(init_categories().keys())
    labels.sort()
    header = ['Model name', 'Total Errors', 'CER in %'] + labels
    print(header)
    rows = [header]
    error_registry = {}
    uncategorized_confusions = []

    for register in global_register.keys():
        error_count = 0
        print("--------------------------------------------")
        print('\033[1m' + register + '\033[0m\n')
        error_numbers = []
        for key in labels:
            print("\t", key, (24 - len(key)) * " ", "\t:", global_register[register][key].count)
            error_numbers.append(int(global_register[register][key].count))
            error_count += global_register[register][key].count
            if key == "Not Categorized":
                uncategorized_confusions.extend(global_register[register][key].confusions)

        error_numbers.insert(0, error_count)
        error_numbers.insert(1, float("{:.2f}".format((error_count/char_count_map[register][0])*100)))
        error_registry[register] = error_numbers
        rows.append([register.replace(".xlsx", "")] + error_numbers)
        print("CER: ", str(error_count/char_count_map[register][0]))
        print("CER: ", str(char_count_map[register][1] / char_count_map[register][0]))
        print("Char Count", str(char_count_map[register][0]))
        print("Error count:", str(error_count))
        print("Error", str(char_count_map[register][1]))
    labels.insert(0, 'Total Errors')
    labels.insert(1, 'CER in %')

    wb = Workbook(write_only=True)
    ws = wb.create_sheet()
    for row in rows:
        print(row)
        ws.append(row)

    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = 'Error Count for multiple models on ' + str(list(char_count_map.values())[0][0]) + ' chars'
    chart1.y_axis.title = 'Error Count'
    chart1.x_axis.title = 'Categories'

    data = Reference(ws, min_col=1, min_row=2, max_row=len(global_register.keys())+1, max_col=len(header))
    cats = Reference(ws, min_row=1, min_col=2, max_col=len(header))
    chart1.add_data(data, titles_from_data=True, from_rows=True)
    chart1.set_categories(cats)
    chart1.shape = 4
    ws.add_chart(chart1, "A" + str(len(global_register.keys()) + 2))
    wb.save(sys.argv[2])

    x = np.arange(len(labels))
    width = 0.5

    fig, ax = plt.subplots()
    rects = []
    count = 0
    for register in error_registry.keys():
        number_list = error_registry[register]
        rect = ax.bar(((x - width/2) + (count * width/len(list(error_registry.keys())))), number_list, width/len(list(error_registry.keys())), label=register)
        rects.append(rect)
        count += 1

    ax.set_ylabel('Error #')
    ax.set_title('Error Count for multiple models on ' + str(list(char_count_map.values())[0][0]) + ' chars')
    ax.set_xticks(x, labels, rotation=45)
    ax.legend()

    for rect in rects:
        ax.bar_label(rect, padding=3, rotation=90)

    unique_confusions = set(uncategorized_confusions)
    for conf in unique_confusions:
        print("{'", conf.gt, "', '", conf.pred, "'}")

    fig.tight_layout()
    plt.show()


def categorize_confusion(categories, confusion):

    if cat_as_whitespace(confusion):
        categories["Whitespace Error"].add(confusion)
    elif cat_as_umlaut(confusion):
        categories["Umlaut Error"].add(confusion)
    elif cat_as_letter_case(confusion):
        categories["Letter Case Error"].add(confusion)
    elif cat_as_resolution(confusion):
        categories["Resolution Error"].add(confusion)
    elif cat_as_unknown_character(confusion):
        categories["Untrained Character"].add(confusion)
    elif cat_as_sub_super(confusion):
        categories["Sub- & Superscript Error"].add(confusion)
    elif cat_as_vert_line(confusion):
        categories["Vertical Line Error"].add(confusion)
    elif cat_as_capitals(confusion):
        categories["Unrecognized Capitals"].add(confusion)
    elif cat_as_kerning(confusion):
        categories["Kerning error"].add(confusion)
    elif cat_as_upper_quote(confusion):
        categories["Upper Quotation Error"].add(confusion)
    elif cat_as_punctuation(confusion):
        categories["Punctuation Error"].add(confusion)
    elif cat_as_similar(confusion):
        categories["Similarity Error"].add(confusion)
    else:
        categories["Not Categorized"].add(confusion)


def init_categories():
    cat_similar = ConfusionCategory("Similarity Error")
    cat_umlaut = ConfusionCategory("Umlaut Error")
    cat_resolution = ConfusionCategory("Resolution Error")
    cat_letter_case = ConfusionCategory("Letter Case Error")
    cat_unknown_char = ConfusionCategory("Untrained Character")
    cat_whitespace = ConfusionCategory("Whitespace Error")
    cat_sub_super = ConfusionCategory("Sub- & Superscript Error")
    cat_vert_line = ConfusionCategory("Vertical Line Error")
    cat_capitals = ConfusionCategory("Unrecognized Capitals")
    cat_kerning = ConfusionCategory("Kerning error")
    cat_upper = ConfusionCategory("Upper Quotation Error")
    cat_punctuation = ConfusionCategory("Punctuation Error")
    cat_unknown_confusion = ConfusionCategory("Not Categorized")
    category_list = {cat_umlaut, cat_letter_case, cat_similar, cat_resolution, cat_unknown_char, cat_whitespace, cat_sub_super, cat_vert_line, cat_capitals, cat_kerning, cat_upper, cat_punctuation, cat_unknown_confusion}
    category_dict = {}
    for cat in category_list:
        category_dict[cat.name] = cat
#    category_dict["Total Chars"] = 0
#    category_dict["Total Errors"] = 0
    return category_dict


def parse_confusions(confusion_string):
    confusions = []
    confusion_string = str(confusion_string.value)[1:-1]
    regex = re.compile(r'(?:[^,(]|\([^)]*\))+')
    confusion_items = regex.findall(confusion_string)
    for item in confusion_items:
        regex2 = re.compile(r'(?:[^:(]|\([^)]*\))+')
        tja = regex2.findall(item)
        if(len(tja) != 2):
            return confusions
        conf_string = tja[0].strip()[1:-1]
        amount = int(tja[1].strip())
        #print(conf_string)
        gt = regex.findall(conf_string)[0][1:-1]
        pred = regex.findall(conf_string)[1].strip()[1:-1]

        confusions.append(Confusion(gt, pred, amount))
    return confusions


def get_char_count(line_sheet):

    count = 0
    err_count = 0
    for row in line_sheet.iter_rows(min_row=2):
        gt_file, gt, pred, line_len, err, cer, rel_err, sync_err, confusions = row
        count += line_len.value
        err_count += err.value
        # err_count += max(err.value, sync_err.value)
        if err.value != sync_err.value:
            print(gt_file.value)
    return count, err_count


if __name__ == "__main__":
    main()